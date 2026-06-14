from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from datetime import date, timedelta
from decimal import Decimal
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db import transaction

from .models import Customer, FollowUp, Order, OrderItem, Product, Payment, Brand
from .services.price_lookup import lookup_product, lookup_local_products, product_to_dict


# ==================== 登录/登出 ====================

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, '用户名或密码错误，请重试')

    return render(request, 'crm/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ==================== 首页仪表盘 ====================

@login_required(login_url='/crm/login/')
def dashboard(request):
    today = date.today()

    total_customers = Customer.objects.count()
    active_customers = Customer.objects.filter(status='active').count()
    potential_customers = Customer.objects.filter(status='potential').count()

    recent_customers = Customer.objects.all()[:5]
    recent_followups = FollowUp.objects.all()[:5]

    birthday_customers = []
    all_customers_with_birthday = Customer.objects.exclude(birthday__isnull=True)

    for customer in all_customers_with_birthday:
        days = customer.days_until_birthday
        if days is not None and 0 <= days <= 7:
            birthday_customers.append({
                'customer': customer,
                'days': days,
                'is_today': days == 0,
            })

    birthday_customers.sort(key=lambda x: x['days'])

    context = {
        'total_customers': total_customers,
        'active_customers': active_customers,
        'potential_customers': potential_customers,
        'recent_customers': recent_customers,
        'recent_followups': recent_followups,
        'birthday_customers': birthday_customers,
        'today': today,
    }
    return render(request, 'crm/dashboard.html', context)


# ==================== 客户管理 ====================

@login_required(login_url='/crm/login/')
def customer_list(request):
    customers = Customer.objects.all()

    search = request.GET.get('search', '')
    if search:
        customers = customers.filter(
            Q(name__icontains=search) |
            Q(institution__icontains=search) |
            Q(lab_group__icontains=search) |
            Q(research_direction__icontains=search) |
            Q(phone__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        customers = customers.filter(status=status)

    source = request.GET.get('source', '')
    if source:
        customers = customers.filter(source=source)

    context = {
        'customers': customers,
        'search': search,
        'status': status,
        'source': source,
    }
    return render(request, 'crm/customer_list.html', context)


@login_required(login_url='/crm/login/')
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    followups = customer.followups.all()

    return render(request, 'crm/customer_detail.html', {
        'customer': customer,
        'followups': followups,
    })


@login_required(login_url='/crm/login/')
def customer_add(request):
    if request.method == 'POST':
        birthday_str = request.POST.get('birthday', '')
        birthday = birthday_str if birthday_str else None

        customer = Customer.objects.create(
            name=request.POST.get('name'),
            institution=request.POST.get('institution', ''),
            lab_group=request.POST.get('lab_group', ''),
            phone=request.POST.get('phone', ''),
            email=request.POST.get('email', ''),
            birthday=birthday,
            research_direction=request.POST.get('research_direction', ''),
            status=request.POST.get('status', 'potential'),
            source=request.POST.get('source', 'other'),
            notes=request.POST.get('notes', ''),
            assigned_to=request.user,
        )
        messages.success(request, f'客户 {customer.name} 已成功添加')
        return redirect('customer_detail', pk=customer.pk)

    return render(request, 'crm/customer_form.html')


@login_required(login_url='/crm/login/')
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)

    if request.method == 'POST':
        birthday_str = request.POST.get('birthday', '')
        customer.name = request.POST.get('name')
        customer.institution = request.POST.get('institution', '')
        customer.lab_group = request.POST.get('lab_group', '')
        customer.phone = request.POST.get('phone', '')
        customer.email = request.POST.get('email', '')
        customer.birthday = birthday_str if birthday_str else None
        customer.research_direction = request.POST.get('research_direction', '')
        customer.status = request.POST.get('status', 'potential')
        customer.source = request.POST.get('source', 'other')
        customer.notes = request.POST.get('notes', '')
        customer.save()
        messages.success(request, f'客户 {customer.name} 信息已更新')
        return redirect('customer_detail', pk=customer.pk)

    return render(request, 'crm/customer_form.html', {'customer': customer})


@login_required(login_url='/crm/login/')
def followup_add(request, customer_pk):
    customer = get_object_or_404(Customer, pk=customer_pk)

    if request.method == 'POST':
        FollowUp.objects.create(
            customer=customer,
            follow_type=request.POST.get('follow_type'),
            content=request.POST.get('content'),
            next_action=request.POST.get('next_action', ''),
            next_date=request.POST.get('next_date') or None,
            created_by=request.user,
        )
        messages.success(request, '跟进记录已保存')
        return redirect('customer_detail', pk=customer_pk)

    return render(request, 'crm/followup_form.html', {'customer': customer})


# ==================== 订单管理 ====================

@login_required(login_url='/crm/login/')
def order_list(request):
    orders = Order.objects.select_related('customer', 'sales_rep').all()

    search = request.GET.get('search', '')
    if search:
        orders = orders.filter(
            Q(order_no__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(customer__institution__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        orders = orders.filter(status=status)

    invoice_status = request.GET.get('invoice_status', '')
    if invoice_status:
        orders = orders.filter(invoice_status=invoice_status)

    payment_status = request.GET.get('payment_status', '')
    if payment_status:
        orders = orders.filter(payment_status=payment_status)

    stats = orders.aggregate(
        total_amount=Sum('final_amount'),
        total_paid=Sum('paid_amount'),
        total_count=Count('id'),
    )

    total_amount = stats['total_amount'] or 0
    total_paid = stats['total_paid'] or 0
    stats['total_unpaid'] = total_amount - total_paid

    context = {
        'orders': orders,
        'search': search,
        'status': status,
        'invoice_status': invoice_status,
        'payment_status': payment_status,
        'stats': stats,
        'status_choices': Order.STATUS_CHOICES,
        'invoice_status_choices': Order.INVOICE_STATUS_CHOICES,
        'payment_status_choices': Order.PAYMENT_STATUS_CHOICES,
    }
    return render(request, 'crm/order_list.html', context)


@login_required(login_url='/crm/login/')
def order_detail(request, pk):
    order = get_object_or_404(
        Order.objects.select_related('customer', 'sales_rep'),
        pk=pk
    )
    items = order.items.select_related('product').all()
    payments = order.payments.all()

    context = {
        'order': order,
        'items': items,
        'payments': payments,
    }
    return render(request, 'crm/order_detail.html', context)


@login_required(login_url='/crm/login/')
def order_create(request):
    if request.method == 'POST':
        try:
            customer = get_object_or_404(Customer, pk=request.POST.get('customer_id'))

            order = Order(
                customer=customer,
                sales_rep=request.user,
                order_date=request.POST.get('order_date') or date.today(),
                status='draft',
                shipping_contact=request.POST.get('shipping_contact', ''),
                shipping_phone=request.POST.get('shipping_phone', ''),
                shipping_address=request.POST.get('shipping_address', ''),
                shipping_method=request.POST.get('shipping_method', 'sf_normal'),
                invoice_type=request.POST.get('invoice_type', 'normal'),
                invoice_title=request.POST.get('invoice_title', ''),
                invoice_tax_id=request.POST.get('invoice_tax_id', ''),
                invoice_email=request.POST.get('invoice_email', ''),
                invoice_remark=request.POST.get('invoice_remark', ''),
                payment_term=request.POST.get('payment_term', 'prepay'),
                shipping_fee=Decimal(request.POST.get('shipping_fee') or '0'),
                discount_amount=Decimal(request.POST.get('discount_amount') or '0'),
                remark=request.POST.get('remark', ''),
            )
            order.save()

            items_json = request.POST.get('items_json', '[]')
            items_data = json.loads(items_json)

            for item_data in items_data:
                if not item_data.get('catalog_number'):
                    continue
                OrderItem.objects.create(
                    order=order,
                    brand_name=item_data.get('brand_name', ''),
                    catalog_number=item_data.get('catalog_number', ''),
                    product_name=item_data.get('product_name', ''),
                    spec=item_data.get('spec', ''),
                    unit=item_data.get('unit', '个'),
                    quantity=Decimal(str(item_data.get('quantity', 1))),
                    list_price=Decimal(str(item_data['list_price']))
                        if item_data.get('list_price') else None,
                    discount=Decimal(str(item_data.get('discount', 100))),
                    unit_price=Decimal(str(item_data.get('unit_price', 0))),
                    cost_price=Decimal(str(item_data['cost_price']))
                        if item_data.get('cost_price') else None,
                    tax_rate=Decimal(str(item_data.get('tax_rate', 13))),
                    remark=item_data.get('remark', ''),
                )

            order.recalculate_amount()

            messages.success(request, f'订单 {order.order_no} 创建成功！')
            return redirect('order_detail', pk=order.pk)

        except Exception as e:
            messages.error(request, f'创建失败：{str(e)}')

    customers = Customer.objects.filter(
        status__in=['active', 'potential']
    ).order_by('name')
    products = Product.objects.filter(
        is_active=True
    ).select_related('brand').order_by('brand__name', 'catalog_number')

    context = {
        'customers': customers,
        'products': products,
        'shipping_method_choices': Order.SHIPPING_METHOD_CHOICES,
        'invoice_type_choices': Order.INVOICE_TYPE_CHOICES,
        'payment_term_choices': Order.PAYMENT_TERM_CHOICES,
        'today': date.today().strftime('%Y-%m-%d'),
    }
    return render(request, 'crm/order_form.html', context)


@login_required(login_url='/crm/login/')
def order_update_status(request, pk):
    order = get_object_or_404(Order, pk=pk)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        tracking = request.POST.get('tracking_number', '').strip()
        invoice_status = request.POST.get('invoice_status', '')

        if new_status and new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            if new_status == 'shipped' and not order.shipped_date:
                order.shipped_date = date.today()

        if tracking:
            order.tracking_number = tracking

        if invoice_status and invoice_status in dict(Order.INVOICE_STATUS_CHOICES):
            order.invoice_status = invoice_status

        order.save()
        messages.success(request, '订单信息已更新')

    return redirect('order_detail', pk=pk)


@login_required(login_url='/crm/login/')
def payment_add(request, order_pk):
    order = get_object_or_404(Order, pk=order_pk)

    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', '0'))
        if amount <= 0:
            messages.error(request, '回款金额必须大于0')
            return redirect('order_detail', pk=order_pk)

        Payment.objects.create(
            order=order,
            amount=amount,
            method=request.POST.get('method', 'transfer'),
            payment_date=request.POST.get('payment_date') or date.today(),
            remark=request.POST.get('remark', ''),
            recorded_by=request.user,
        )
        messages.success(request, f'回款记录已添加')
        return redirect('order_detail', pk=order_pk)

    return redirect('order_detail', pk=order_pk)


@login_required(login_url='/crm/login/')
def order_bulk_export(request):
    if request.method != 'POST':
        return redirect('order_list')

    order_ids = request.POST.getlist('order_ids')
    action = request.POST.get('action', 'export')

    if not order_ids:
        messages.warning(request, '未选择任何订单')
        return redirect('order_list')

    orders = Order.objects.filter(
        pk__in=order_ids
    ).select_related('customer', 'sales_rep').prefetch_related('items')

    if action == 'mark_submitted':
        updated = orders.exclude(invoice_type='none').update(
            invoice_status='submitted'
        )
        messages.success(request, f'已将 {updated} 个订单标记为【已递交财务】')
        return redirect('order_list')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '开票明细'

    header_font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(
        start_color='2E75B6', end_color='2E75B6', fill_type='solid'
    )
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        '订单编号', '订单日期', '客户姓名', '单位',
        '发票抬头', '税号', '发票类型',
        '品牌', '货号', '产品名称', '规格', '单位',
        '数量', '单价', '小计',
        '运费', '优惠', '订单总额',
        '发货方式', '快递单号',
        '付款方式', '付款状态', '备注',
    ]
    col_widths = [
        18, 12, 10, 20,
        25, 20, 12,
        10, 15, 30, 10, 6,
        8, 10, 10,
        8, 8, 12,
        12, 15,
        10, 10, 15,
    ]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = 'A2'
    row_num = 2

    for order in orders:
        items = list(order.items.all()) or [None]
        for i, item in enumerate(items):
            if i == 0:
                base = [
                    order.order_no,
                    order.order_date.strftime('%Y-%m-%d'),
                    order.customer.name,
                    order.customer.institution,
                    order.invoice_title,
                    order.invoice_tax_id,
                    order.get_invoice_type_display(),
                ]
            else:
                base = ['', '', '', '', '', '', '']

            if item:
                prod = [
                    item.brand_name,
                    item.catalog_number,
                    item.product_name,
                    item.spec,
                    item.unit,
                    float(item.quantity),
                    float(item.unit_price),
                    float(item.subtotal),
                ]
            else:
                prod = ['', '', '', '', '', '', '', '']

            if i == 0:
                tail = [
                    float(order.shipping_fee),
                    float(order.discount_amount),
                    float(order.final_amount),
                    order.get_shipping_method_display(),
                    order.tracking_number,
                    order.get_payment_term_display(),
                    dict(Order.PAYMENT_STATUS_CHOICES).get(
                        order.payment_status, ''
                    ),
                    order.remark,
                ]
            else:
                tail = ['', '', '', '', '', '', '', '']

            row_data = base + prod + tail
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin
                cell.alignment = left

            row_num += 1

    today_str = date.today().strftime('%Y%m%d')
    filename = '开票明细_{}.xlsx'.format(today_str)
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    wb.save(response)
    return response


@login_required(login_url='/crm/login/')
def export_orders_excel(request):
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    export_type = request.GET.get('export_type', 'all')

    orders = Order.objects.select_related(
        'customer', 'sales_rep'
    ).prefetch_related('items').exclude(status='cancelled')

    if start_date:
        orders = orders.filter(order_date__gte=start_date)
    if end_date:
        orders = orders.filter(order_date__lte=end_date)

    if export_type == 'invoice':
        orders = orders.exclude(invoice_type='none')
    elif export_type == 'no_invoice':
        orders = orders.filter(invoice_type='none')

    wb = openpyxl.Workbook()
    header_font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(
        start_color='2E75B6', end_color='2E75B6', fill_type='solid'
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    def write_orders_to_sheet(ws, orders_qs, sheet_title):
        ws.title = sheet_title
        headers = [
            '订单编号', '订单日期', '客户姓名', '所在单位', '课题组',
            '业务员', '订单状态',
            '品牌', '货号', '产品名称', '规格', '单位', '数量', '单价', '小计',
            '运费', '优惠金额', '订单总额', '已付金额', '待付金额',
            '发票类型', '开票状态', '发票抬头', '税号', '发票邮箱',
            '发货方式', '快递单号', '发货日期',
            '付款方式', '结算方式',
            '备注',
        ]
        col_widths = [
            18, 12, 10, 25, 15,
            8, 10,
            12, 15, 30, 12, 6, 8, 10, 10,
            8, 8, 10, 10, 10,
            12, 10, 25, 20, 20,
            12, 15, 12,
            10, 10,
            20,
        ]

        for col_idx, (header, width) in enumerate(
            zip(headers, col_widths), start=1
        ):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

        row_num = 2
        for order in orders_qs:
            items = list(order.items.all())
            if not items:
                items = [None]

            for i, item in enumerate(items):
                if i == 0:
                    order_info = [
                        order.order_no,
                        order.order_date.strftime('%Y-%m-%d'),
                        order.customer.name,
                        order.customer.institution,
                        order.customer.lab_group,
                        order.sales_rep.get_full_name() or order.sales_rep.username,
                        order.get_status_display(),
                    ]
                else:
                    order_info = ['', '', '', '', '', '', '']

                if item:
                    item_info = [
                        item.brand_name,
                        item.catalog_number,
                        item.product_name,
                        item.spec,
                        item.unit,
                        float(item.quantity),
                        float(item.unit_price),
                        float(item.subtotal),
                    ]
                else:
                    item_info = ['', '', '', '', '', '', '', '']

                if i == 0:
                    amount_info = [
                        float(order.shipping_fee),
                        float(order.discount_amount),
                        float(order.final_amount),
                        float(order.paid_amount),
                        float(order.unpaid_amount),
                    ]
                    invoice_info = [
                        order.get_invoice_type_display(),
                        order.get_invoice_status_display(),
                        order.invoice_title,
                        order.invoice_tax_id,
                        order.invoice_email,
                    ]
                    shipping_info = [
                        order.get_shipping_method_display(),
                        order.tracking_number,
                        order.shipped_date.strftime('%Y-%m-%d')
                            if order.shipped_date else '',
                    ]
                    payment_info = [
                        order.get_payment_term_display(),
                        order.get_payment_method_display(),
                    ]
                    remark_info = [order.remark]
                else:
                    amount_info = ['', '', '', '', '']
                    invoice_info = ['', '', '', '', '']
                    shipping_info = ['', '', '']
                    payment_info = ['', '']
                    remark_info = ['']

                row_data = (
                    order_info + item_info + amount_info +
                    invoice_info + shipping_info + payment_info + remark_info
                )

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = left_align
                    if col_idx in [14, 15, 17, 18, 19, 20]:
                        cell.number_format = '#,##0.00'

                row_num += 1

    ws_all = wb.active
    need_invoice_orders = orders.exclude(invoice_type='none')
    no_invoice_orders = orders.filter(invoice_type='none')

    write_orders_to_sheet(ws_all, orders, '全部订单')

    ws_invoice = wb.create_sheet()
    write_orders_to_sheet(ws_invoice, need_invoice_orders, '需开票订单')

    ws_no_invoice = wb.create_sheet()
    write_orders_to_sheet(ws_no_invoice, no_invoice_orders, '不开票订单')

    today_str = date.today().strftime('%Y%m%d')
    filename = '订单导出_{}.xlsx'.format(today_str)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    wb.save(response)
    return response


# ==================== 产品管理 ====================

@login_required(login_url='/crm/login/')
def product_list(request):
    products = Product.objects.select_related('brand').all()

    search = request.GET.get('search', '').strip()
    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(catalog_number__icontains=search) |
            Q(name_en__icontains=search) |
            Q(cas_number__icontains=search)
        )

    brand_id = request.GET.get('brand', '')
    if brand_id:
        products = products.filter(brand_id=brand_id)

    category = request.GET.get('category', '')
    if category:
        products = products.filter(category=category)

    is_active = request.GET.get('is_active', '')
    if is_active == '1':
        products = products.filter(is_active=True)
    elif is_active == '0':
        products = products.filter(is_active=False)

    brands = Brand.objects.all()

    context = {
        'products': products[:200],
        'total_count': products.count(),
        'brands': brands,
        'search': search,
        'brand_id': brand_id,
        'category': category,
        'is_active': is_active,
        'category_choices': Product.CATEGORY_CHOICES,
    }
    return render(request, 'crm/product_list.html', context)


@login_required(login_url='/crm/login/')
def product_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        brand_id = request.POST.get('brand_id')

        if not excel_file:
            messages.error(request, '请选择文件')
            return redirect('product_import')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            headers = []
            for cell in ws[1]:
                val = cell.value
                headers.append(str(val).strip() if val is not None else '')

            keyword_map = {
                'catalog_number': [
                    '货号', 'cat no', 'cat.no', 'catalog',
                    '产品编号', 'item no', 'sku', 'cat#'
                ],
                'name': [
                    '产品名称', 'product name', '品名',
                    'description', '名称', '中文名'
                ],
                'name_en': ['英文名', 'english name', 'english'],
                'spec': [
                    '规格', 'size', 'package', '包装',
                    '规格型号', 'specification'
                ],
                'unit': ['单位', 'unit', '包装单位'],
                'list_price': [
                    '目录价', 'list price', '官网价',
                    '含税价', '价格', 'price'
                ],
                'dealer_price': [
                    '经销商价', '经销商', '进价', 'cost',
                    '成本价', '代理价', '折扣价'
                ],
                'terminal_price': [
                    '终端价', '终端', '售价', '销售价',
                    'sell price', '零售价'
                ],
                'cas_number': ['cas', 'cas号', 'cas no', 'cas#'],
                'storage': ['存储', '保存条件', 'storage', '储存'],
                'lead_time': ['货期', 'lead time', '交货期'],
                'description': ['描述', '说明', '备注'],
            }

            col_map = {}
            for field, keywords in keyword_map.items():
                for i, header in enumerate(headers):
                    header_lower = header.lower().strip()
                    if any(kw in header_lower for kw in keywords):
                        col_map[field] = i
                        break

            if 'catalog_number' not in col_map:
                messages.error(
                    request,
                    '未找到货号列！当前识别到的表头：'
                    + '、'.join(h for h in headers if h)
                )
                return redirect('product_import')

            brand = None
            if brand_id:
                try:
                    brand = Brand.objects.get(pk=brand_id)
                except Brand.DoesNotExist:
                    pass

            success_count = 0
            update_count = 0
            skip_count = 0
            error_rows = []
            skip_rows = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                values = [cell.value for cell in row]

                if all(v is None for v in values):
                    skip_count += 1
                    continue

                catalog_col = col_map.get('catalog_number', 0)
                catalog_val = (
                    values[catalog_col]
                    if catalog_col < len(values)
                    else None
                )

                if not catalog_val:
                    skip_rows.append({
                        'row': row_idx,
                        'reason': '货号为空',
                        'data': str(values[:5]),
                    })
                    skip_count += 1
                    continue

                catalog_number = str(catalog_val).strip()

                def get_str(field, default=''):
                    col = col_map.get(field)
                    if col is not None and col < len(values):
                        v = values[col]
                        if v is not None:
                            return str(v).strip()
                    return default

                def get_decimal(field):
                    col = col_map.get(field)
                    if col is not None and col < len(values):
                        v = values[col]
                        if v is not None and str(v).strip():
                            try:
                                clean = (
                                    str(v)
                                    .replace('¥', '')
                                    .replace(',', '')
                                    .replace(' ', '')
                                    .strip()
                                )
                                if clean:
                                    return Decimal(clean)
                            except Exception:
                                pass
                    return None

                def get_int(field, default=7):
                    col = col_map.get(field)
                    if col is not None and col < len(values):
                        v = values[col]
                        if v is not None:
                            try:
                                return int(float(str(v)))
                            except Exception:
                                pass
                    return default

                try:
                    defaults = {
                        'name': get_str('name') or catalog_number,
                        'name_en': get_str('name_en'),
                        'spec': get_str('spec'),
                        'unit': get_str('unit') or '个',
                        'cas_number': get_str('cas_number'),
                        'description': get_str('description'),
                        'lead_time': get_int('lead_time', 7),
                        'list_price': get_decimal('list_price'),
                        'dealer_price': get_decimal('dealer_price'),
                        'terminal_price': get_decimal('terminal_price'),
                        'is_active': True,
                    }

                    if brand:
                        defaults['brand'] = brand

                    product, created = Product.objects.update_or_create(
                        catalog_number=catalog_number,
                        defaults=defaults
                    )

                    if created:
                        success_count += 1
                    else:
                        update_count += 1

                except Exception as e:
                    error_rows.append({
                        'row': row_idx,
                        'catalog': catalog_number,
                        'reason': str(e),
                        'data': str(values[:8]),
                    })

            total_msg = (
                '导入完成：新增 {} 条，更新 {} 条，跳过 {} 条'.format(
                    success_count, update_count, skip_count
                )
            )
            if error_rows:
                total_msg += '，失败 {} 条'.format(len(error_rows))

            if success_count + update_count > 0:
                messages.success(request, total_msg)
            else:
                messages.warning(request, total_msg)

            request.session['import_errors'] = error_rows[:50]
            request.session['import_skips'] = skip_rows[:20]
            request.session['import_col_map'] = {
                k: headers[v] for k, v in col_map.items()
            }

        except Exception as e:
            messages.error(request, '文件解析失败：{}'.format(str(e)))
            request.session['import_errors'] = []

        return redirect('product_import')

    brands = Brand.objects.all().order_by('name')
    import_errors = request.session.pop('import_errors', [])
    import_skips = request.session.pop('import_skips', [])
    import_col_map = request.session.pop('import_col_map', {})

    return render(request, 'crm/product_import.html', {
        'brands': brands,
        'import_errors': import_errors,
        'import_skips': import_skips,
        'import_col_map': import_col_map,
    })


@login_required(login_url='/crm/login/')
def product_inline_edit(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False}, status=405)

    product = get_object_or_404(Product, pk=pk)

    try:
        field = request.POST.get('field')
        value = request.POST.get('value', '').strip()

        allowed_fields = [
            'name', 'name_en', 'catalog_number', 'spec',
            'unit', 'dealer_price', 'terminal_price',
            'list_price', 'lead_time', 'is_active',
            'storage', 'cas_number', 'description',
        ]

        if field not in allowed_fields:
            return JsonResponse({'success': False, 'error': '不允许编辑此字段'})

        decimal_fields = ['dealer_price', 'terminal_price', 'list_price']
        int_fields = ['lead_time']
        bool_fields = ['is_active']

        if field in decimal_fields:
            value = Decimal(value) if value else None
        elif field in int_fields:
            value = int(value) if value else 0
        elif field in bool_fields:
            value = value.lower() in ('true', '1', 'yes')

        setattr(product, field, value)
        product.save(update_fields=[field])

        return JsonResponse({
            'success': True,
            'value': str(getattr(product, field)),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required(login_url='/crm/login/')
def product_bulk_delete(request):
    if request.method == 'POST':
        ids = request.POST.getlist('product_ids')
        if ids:
            count = Product.objects.filter(pk__in=ids).count()
            Product.objects.filter(pk__in=ids).delete()
            messages.success(request, '已删除 {} 个产品'.format(count))
        else:
            messages.warning(request, '未选择任何产品')
    return redirect('product_list')


# ==================== API ====================

@login_required(login_url='/crm/login/')
def product_search_api(request):
    q = request.GET.get('q', '').strip()
    customer_id = request.GET.get('customer_id', '')

    if len(q) < 1:
        return JsonResponse({'products': []})

    products = Product.objects.filter(
        Q(name__icontains=q) |
        Q(catalog_number__icontains=q) |
        Q(name_en__icontains=q),
        is_active=True
    ).select_related('brand')[:20]

    # 获取客户折扣信息
    customer = None
    if customer_id:
        try:
            customer = Customer.objects.get(pk=customer_id)
        except Customer.DoesNotExist:
            pass

    data = []
    for p in products:
        # 计算该客户对该品牌的折扣
        discount = Decimal('100')
        if customer:
            discount = customer.get_discount_for_brand(p.brand)

        list_price = float(p.terminal_price) if p.terminal_price else 0
        discounted_price = round(list_price * float(discount) / 100, 2)

        data.append({
            'id': p.id,
            'brand_name': p.brand.name if p.brand else '',
            'brand_id': p.brand.id if p.brand else None,
            'catalog_number': p.catalog_number,
            'product_name': p.name,
            'spec': p.spec,
            'unit': p.unit,
            'terminal_price': str(p.terminal_price) if p.terminal_price else '',
            'dealer_price': str(p.dealer_price) if p.dealer_price else '',
            'discount': str(discount),
            'discounted_price': str(discounted_price) if list_price > 0 else '',
            'display': str(p),
        })

    return JsonResponse({'products': data})

@login_required(login_url='/crm/login/')
def customer_info_api(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    last = customer.get_last_order_info()

    # 获取该客户的默认折扣（不指定品牌）
    default_discount = customer.get_discount_for_brand(None)

    data = {
        'invoice_title': customer.invoice_title or last.get('invoice_title', ''),
        'invoice_tax_id': customer.invoice_tax_id or last.get('invoice_tax_id', ''),
        'invoice_bank': customer.invoice_bank or last.get('invoice_bank', ''),
        'invoice_bank_account': (
            customer.invoice_bank_account
            or last.get('invoice_bank_account', '')
        ),
        'invoice_address': (
            customer.invoice_address or last.get('invoice_address', '')
        ),
        'invoice_phone': (
            customer.invoice_phone or last.get('invoice_phone', '')
        ),
        'invoice_email': customer.email or last.get('invoice_email', ''),
        'invoice_type': last.get('invoice_type', 'normal'),
        'shipping_contact': (
            customer.default_shipping_contact
            or last.get('shipping_contact', '')
        ),
        'shipping_phone': (
            customer.default_shipping_phone
            or last.get('shipping_phone', '')
        ),
        'shipping_address': (
            customer.default_shipping_address
            or last.get('shipping_address', '')
        ),
        'shipping_method': last.get('shipping_method', 'sf_normal'),
        'payment_term': last.get('payment_term', 'prepay'),
        'payment_method': last.get('payment_method', 'transfer'),
        # 折扣信息
        'discount_rate': str(default_discount),
        'customer_type': customer.customer_type,
        'customer_type_display': customer.get_customer_type_display(),
        'level': customer.level,
        'discount_text': customer.get_discount_display_text(),
    }
    return JsonResponse(data)


@login_required(login_url='/crm/login/')
def brand_autocomplete_api(request):
    """
    品牌名称自动完成API
    返回匹配的品牌列表，支持中英文名称搜索
    """
    q = request.GET.get('q', '').strip()

    if len(q) < 1:
        return JsonResponse({'brands': []})

    brands = Brand.objects.filter(
        Q(name__icontains=q) |
        Q(name_en__icontains=q)
    ).order_by('name')[:20]

    data = []
    for brand in brands:
        data.append({
            'id': brand.id,
            'name': brand.name,
            'name_en': brand.name_en,
            'country': brand.country,
            'display': f"{brand.name} ({brand.name_en})" if brand.name_en else brand.name,
            'has_template': bool(brand.search_url_template),
        })

    return JsonResponse({'brands': data, 'has_other': True})


@login_required(login_url='/crm/login/')
def product_lookup_api(request):
    """
    产品查价API
    根据品牌ID + 货号查询产品，结合本地库和官网价格
    """
    from .services.price_lookup import lookup_product

    brand_id = request.GET.get('brand_id', '')
    catalog_number = request.GET.get('catalog_number', '').strip()
    customer_id = request.GET.get('customer_id', '')

    if not catalog_number:
        return JsonResponse({
            'success': False,
            'error': '请输入货号',
        })

    # 获取客户对象（用于计算折扣）
    customer = None
    if customer_id:
        try:
            customer = Customer.objects.get(pk=customer_id)
        except Customer.DoesNotExist:
            pass

    # 调用价格查询服务
    result = lookup_product(
        brand_id=brand_id if brand_id else None,
        catalog_number=catalog_number,
        customer=customer,
    )

    # 构建响应
    data = {
        'success': True,
        'catalog_number': result['catalog_number'],
        'brand_name': result['brand_name'],
        'brand_id': result['brand_id'],
        'local_products': result['local_products'],
        'has_local': result['has_local'],
        'website_url': result['website_url'],
        'website_price': result['website_price'],
        'price_source': result['price_source'],
    }

    # 如果有本地产品，返回最佳匹配
    if result['local_products']:
        best = result['local_products'][0]
        data['product'] = {
            'id': best['id'],
            'name': best['product_name'],
            'name_en': best.get('name_en', ''),
            'spec': best['spec'],
            'unit': best['unit'],
            'list_price': best['list_price'],
            'terminal_price': best['terminal_price'],
            'dealer_price': best['dealer_price'],
            'discount': best['discount'],
            'discounted_price': best['discounted_price'],
        }
    elif result['website_price']:
        # 仅有官网价格
        data['product'] = {
            'id': None,
            'name': '',
            'name_en': '',
            'spec': '',
            'unit': '个',
            'list_price': result['website_price'],
            'terminal_price': result['website_price'],
            'dealer_price': '',
            'discount': '100',
            'discounted_price': result['website_price'],
        }

    return JsonResponse(data)
